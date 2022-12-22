import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet



def check_file(file_name: str) -> bool:
    """
    SUMMARY
    -------
    This function check if the file passed as parameter exists in the assets folder.

    ARGUMENTS
    ---------
        - file_name : str
            The name of the file that we want check.

    RETURNS
    -------
    bool: True if the file exists in the assets folder on False else.
    """
    return os.path.exists(os.path.join(os.path.realpath(os.path.dirname(__file__)), "..", "..", "assets", file_name))


def get_cell_data(excel_file_name: str, letter_column: str, row: int) -> str:
    """
    SUMMARY
    -------
    This function get the data stored in the excel file to the specific cell at the coordinate passed as parameter.

    ARGUMENTS
    ---------
        - excel_file_name : str
            The name of the excel file.
        - letter_column : str
            The letter of the cell column in the excel file serving as abscissa.
        - row : int
            The row of the cell column in the excel file serving as ordinate.

    RETURNS
    -------
    str: The value content of the cell.
    """

    # load the excel file
    wb = load_workbook(os.path.join(os.path.realpath(os.path.dirname(__file__)), "..", "..", "assets", excel_file_name))
    ws = wb.active

    # get the cell
    cell_data = ws[letter_column + str(row)]

    # close the excel file and the value of the cell
    wb.close()
    return cell_data.value


def get_row_data(excel_file_name: str, parameters: dict[str, tuple[str, str]], row: int) -> dict[str, str]:
    """
    SUMMARY
    -------
    This function get the data in the excel file to all parameters passed as arguments on a specific row.

    ARGUMENTS
    ---------
        - excel_file_name : str
            The name of the excel file.
        - parameters : dict[str, tuple[str, str]]
            The list of parameters in the format (parameter name => (default value, letter column)).
        - row : int
            The row of the cell column in the excel file serving as ordinate.

    RETURNS
    -------
    dict[str, str]: The list of the parameters with their value corresponding.
    """

    # load the excel file
    wb = load_workbook(os.path.join(os.path.realpath(os.path.dirname(__file__)), "..", "..", "assets", excel_file_name))
    ws = wb.active

    # build the list of the parameters with their value corresponding
    params_value = dict()
    for param_name, values in parameters.items():
        # excel cell content
        current_data = ws[values[1] + str(row)].value

        if current_data is None:
            params_value[param_name] = values[0] # default parameter value
        else:
            params_value[param_name] = current_data

    # close the excel file and return the parameters and their value
    wb.close()
    return params_value


def check_duplicates(excel_file_name: str, letter_column: str, start_row: str, end_row: str) -> int:
    wb = load_workbook(os.path.join(os.path.realpath(os.path.dirname(__file__)), "..", "..", "assets", excel_file_name))
    ws = wb.active

    cells_data = dict()
    for row in range(start_row, end_row + 1):
        current_cell_data = ws[letter_column + str(row)].value

        if current_cell_data in cells_data.values():
            ws[letter_column + str(__get_key_by_value(cells_data, current_cell_data))].fill = PatternFill(start_color="32CD32", fill_type="solid")
            ws[letter_column + str(row)].fill = PatternFill(start_color="DC143C", fill_type="solid")

        else:
            cells_data[row] = current_cell_data

    wb.save(os.path.join(os.path.realpath(os.path.dirname(__file__)), "..", "..", "assets", excel_file_name))
    wb.close()
    return end_row + 1 - start_row - len(cells_data)


def __get_key_by_value(dictionary: dict[int, str], value: str) -> int or None:
    for key, current_value in dictionary.items():
        if value == current_value:
            return key
    
    return None
